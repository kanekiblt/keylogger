from openpyxl import Workbook, load_workbook
from datetime import datetime
import pygetwindow as gw
import os

ARCHIVO_WSP = "wsp_mensajes.xlsx"
KEYWORDS_WSP = ["whatsapp"]

chat_actual = "Chat_1"

if not os.path.exists(ARCHIVO_WSP):
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Perfil", "Mensaje", "Link"])
    wb.save(ARCHIVO_WSP)

def es_whatsapp_activo():
    try:
        ventana = gw.getActiveWindow()
        if ventana:
            titulo = ventana.title.lower()
            return any(k in titulo for k in KEYWORDS_WSP)
    except:
        pass
    return False

def obtener_perfil():
    global chat_actual

    try:
        ventana = gw.getActiveWindow()
        if ventana:
            titulo = ventana.title.strip()

            if "-" in titulo:
                perfil = titulo.split("-")[0].strip()

                if perfil.lower() != "whatsapp" and perfil != "":
                    chat_actual = perfil
                    return perfil

            return chat_actual
    except:
        pass

    return chat_actual

def guardar_mensaje(mensaje):
    try:
        wb = load_workbook(ARCHIVO_WSP)
        ws = wb.active

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        perfil = obtener_perfil()

        ws.append([fecha, perfil, mensaje, "https://web.whatsapp.com/"])
        wb.save(ARCHIVO_WSP)
    except:
        pass