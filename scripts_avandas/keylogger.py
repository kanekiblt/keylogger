from openpyxl import Workbook, load_workbook
import os

ARCHIVO_KEYS = "contador_teclas.xlsx"

if not os.path.exists(ARCHIVO_KEYS):
    wb = Workbook()
    ws = wb.active
    ws.append(["Tecla", "Cantidad"])
    wb.save(ARCHIVO_KEYS)

def contar_tecla(tecla):
    try:
        wb = load_workbook(ARCHIVO_KEYS)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if row[0].value == tecla:
                row[1].value += 1
                wb.save(ARCHIVO_KEYS)
                return

        ws.append([tecla, 1])
        wb.save(ARCHIVO_KEYS)

    except:
        pass